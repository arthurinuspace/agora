"""
Poll export utilities for Agora Slack app.
Handles exporting poll results to various formats (CSV, JSON, Excel).
"""

import csv
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, IO
from io import StringIO, BytesIO
from dataclasses import dataclass
from sqlalchemy.orm import Session
from models import Poll, PollOption, UserVote, VotedUser, SessionLocal
from performance import OptimizedQueries

logger = logging.getLogger(__name__)

@dataclass
class ExportOptions:
    """Export configuration options."""
    include_voter_ids: bool = False
    include_timestamps: bool = True
    include_analytics: bool = True
    include_metadata: bool = True
    group_by_option: bool = False
    anonymize_data: bool = True

class PollExporter:
    """Handles poll data export to various formats."""
    
    def __init__(self):
        self.supported_formats = ['csv', 'json', 'excel']
    
    def export_poll(self, poll_id: int, format_type: str, options: ExportOptions = None) -> Optional[bytes]:
        """Export a single poll to specified format."""
        if format_type not in self.supported_formats:
            logger.error(f"Unsupported export format: {format_type}")
            return None
        
        options = options or ExportOptions()
        
        try:
            # Get poll data
            poll_data = self._get_poll_export_data(poll_id, options)
            if not poll_data:
                logger.error(f"No data found for poll {poll_id}")
                return None
            
            # Export based on format
            if format_type == 'csv':
                return self._export_to_csv(poll_data, options)
            elif format_type == 'json':
                return self._export_to_json(poll_data, options)
            elif format_type == 'excel':
                return self._export_to_excel(poll_data, options)
        
        except Exception as e:
            logger.error(f"Error exporting poll {poll_id}: {e}")
            return None
    
    def export_multiple_polls(self, poll_ids: List[int], format_type: str, options: ExportOptions = None) -> Optional[bytes]:
        """Export multiple polls to specified format."""
        if format_type not in self.supported_formats:
            logger.error(f"Unsupported export format: {format_type}")
            return None
        
        options = options or ExportOptions()
        
        try:
            # Get data for all polls
            all_poll_data = []
            for poll_id in poll_ids:
                poll_data = self._get_poll_export_data(poll_id, options)
                if poll_data:
                    all_poll_data.append(poll_data)
            
            if not all_poll_data:
                logger.error("No data found for any of the specified polls")
                return None
            
            # Export based on format
            if format_type == 'csv':
                return self._export_multiple_to_csv(all_poll_data, options)
            elif format_type == 'json':
                return self._export_multiple_to_json(all_poll_data, options)
            elif format_type == 'excel':
                return self._export_multiple_to_excel(all_poll_data, options)
        
        except Exception as e:
            logger.error(f"Error exporting multiple polls: {e}")
            return None
    
    def _get_poll_export_data(self, poll_id: int, options: ExportOptions) -> Optional[Dict[str, Any]]:
        """Get poll data for export."""
        try:
            db = SessionLocal()
            
            # Get poll with details
            poll = OptimizedQueries.get_poll_with_details(db, poll_id)
            if not poll:
                return None
            
            # Get analytics if requested
            analytics = None
            if options.include_analytics:
                analytics = OptimizedQueries.get_poll_analytics(db, poll_id)
            
            # Get votes data
            votes_data = []
            if options.include_voter_ids and not options.anonymize_data:
                # Include voter information (admin only)
                votes = db.query(UserVote).filter(UserVote.poll_id == poll_id).all()
                for vote in votes:
                    vote_data = {
                        'option_id': vote.option_id,
                        'user_id': vote.user_id if not options.anonymize_data else f"User_{hash(vote.user_id) % 10000}",
                        'voted_at': vote.voted_at.isoformat() if options.include_timestamps and vote.voted_at else None
                    }
                    votes_data.append(vote_data)
            else:
                # Anonymous vote counts only
                for option in poll.options:
                    votes_data.append({
                        'option_id': option.id,
                        'option_text': option.text,
                        'vote_count': option.vote_count,
                        'order_index': option.order_index
                    })
            
            # Build export data
            export_data = {
                'poll_id': poll.id,
                'question': poll.question,
                'vote_type': poll.vote_type,
                'status': poll.status,
                'options': [
                    {
                        'id': opt.id,
                        'text': opt.text,
                        'vote_count': opt.vote_count,
                        'order_index': opt.order_index
                    }
                    for opt in poll.options
                ],
                'votes': votes_data
            }
            
            # Add metadata if requested
            if options.include_metadata:
                export_data['metadata'] = {
                    'team_id': poll.team_id,
                    'channel_id': poll.channel_id,
                    'creator_id': poll.creator_id if not options.anonymize_data else "Anonymous",
                    'created_at': poll.created_at.isoformat() if options.include_timestamps else None,
                    'ended_at': poll.ended_at.isoformat() if poll.ended_at and options.include_timestamps else None,
                    'message_ts': poll.message_ts
                }
            
            # Add analytics if requested
            if options.include_analytics and analytics:
                export_data['analytics'] = analytics
            
            return export_data
        
        except Exception as e:
            logger.error(f"Error getting poll export data: {e}")
            return None
        
        finally:
            db.close()
    
    def _export_to_csv(self, poll_data: Dict[str, Any], options: ExportOptions) -> bytes:
        """Export poll data to CSV format."""
        output = StringIO()
        
        if options.group_by_option:
            # Group votes by option
            writer = csv.writer(output)
            
            # Header
            headers = ['Poll ID', 'Question', 'Option', 'Vote Count']
            if options.include_metadata:
                headers.extend(['Team ID', 'Channel ID', 'Creator', 'Created At'])
            writer.writerow(headers)
            
            # Data rows
            for option in poll_data['options']:
                row = [
                    poll_data['poll_id'],
                    poll_data['question'],
                    option['text'],
                    option['vote_count']
                ]
                
                if options.include_metadata:
                    metadata = poll_data.get('metadata', {})
                    row.extend([
                        metadata.get('team_id', ''),
                        metadata.get('channel_id', ''),
                        metadata.get('creator_id', ''),
                        metadata.get('created_at', '')
                    ])
                
                writer.writerow(row)
        
        else:
            # Individual votes (if available)
            writer = csv.writer(output)
            
            # Header
            headers = ['Poll ID', 'Question', 'Option ID', 'Option Text']
            if options.include_voter_ids and not options.anonymize_data:
                headers.append('User ID')
            if options.include_timestamps:
                headers.append('Voted At')
            writer.writerow(headers)
            
            # Data rows
            for vote in poll_data['votes']:
                if 'option_text' in vote:  # Anonymous vote counts
                    row = [
                        poll_data['poll_id'],
                        poll_data['question'],
                        vote['option_id'],
                        vote['option_text']
                    ]
                else:  # Individual votes
                    option = next((opt for opt in poll_data['options'] if opt['id'] == vote['option_id']), {})
                    row = [
                        poll_data['poll_id'],
                        poll_data['question'],
                        vote['option_id'],
                        option.get('text', '')
                    ]
                    
                    if options.include_voter_ids and not options.anonymize_data:
                        row.append(vote.get('user_id', ''))
                    if options.include_timestamps:
                        row.append(vote.get('voted_at', ''))
                
                writer.writerow(row)
        
        return output.getvalue().encode('utf-8')
    
    def _export_to_json(self, poll_data: Dict[str, Any], options: ExportOptions) -> bytes:
        """Export poll data to JSON format."""
        # Add export metadata
        export_data = {
            'exported_at': datetime.now().isoformat(),
            'export_options': {
                'include_voter_ids': options.include_voter_ids,
                'include_timestamps': options.include_timestamps,
                'include_analytics': options.include_analytics,
                'include_metadata': options.include_metadata,
                'anonymize_data': options.anonymize_data
            },
            'poll_data': poll_data
        }
        
        return json.dumps(export_data, indent=2, ensure_ascii=False).encode('utf-8')
    
    def _export_to_excel(self, poll_data: Dict[str, Any], options: ExportOptions) -> bytes:
        """Export poll data to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            workbook = openpyxl.Workbook()
            
            # Summary sheet
            summary_sheet = workbook.active
            summary_sheet.title = "Poll Summary"
            
            # Headers
            summary_sheet['A1'] = "Poll Summary"
            summary_sheet['A1'].font = Font(bold=True, size=14)
            
            # Poll info
            row = 3
            summary_sheet[f'A{row}'] = "Poll ID:"
            summary_sheet[f'B{row}'] = poll_data['poll_id']
            row += 1
            
            summary_sheet[f'A{row}'] = "Question:"
            summary_sheet[f'B{row}'] = poll_data['question']
            row += 1
            
            summary_sheet[f'A{row}'] = "Vote Type:"
            summary_sheet[f'B{row}'] = poll_data['vote_type']
            row += 1
            
            summary_sheet[f'A{row}'] = "Status:"
            summary_sheet[f'B{row}'] = poll_data['status']
            row += 2
            
            # Options and results
            summary_sheet[f'A{row}'] = "Option"
            summary_sheet[f'B{row}'] = "Vote Count"
            summary_sheet[f'A{row}'].font = Font(bold=True)
            summary_sheet[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for option in poll_data['options']:
                summary_sheet[f'A{row}'] = option['text']
                summary_sheet[f'B{row}'] = option['vote_count']
                row += 1
            
            # Add metadata sheet if requested
            if options.include_metadata and 'metadata' in poll_data:
                metadata_sheet = workbook.create_sheet("Metadata")
                metadata = poll_data['metadata']
                
                metadata_sheet['A1'] = "Metadata"
                metadata_sheet['A1'].font = Font(bold=True, size=14)
                
                row = 3
                for key, value in metadata.items():
                    metadata_sheet[f'A{row}'] = key.replace('_', ' ').title()
                    metadata_sheet[f'B{row}'] = str(value) if value else ''
                    row += 1
            
            # Add analytics sheet if requested
            if options.include_analytics and 'analytics' in poll_data:
                analytics_sheet = workbook.create_sheet("Analytics")
                analytics = poll_data['analytics']
                
                analytics_sheet['A1'] = "Analytics"
                analytics_sheet['A1'].font = Font(bold=True, size=14)
                
                row = 3
                for key, value in analytics.items():
                    if key != 'vote_distribution':  # Handle separately
                        analytics_sheet[f'A{row}'] = key.replace('_', ' ').title()
                        analytics_sheet[f'B{row}'] = str(value) if value else ''
                        row += 1
            
            # Save to bytes
            output = BytesIO()
            workbook.save(output)
            return output.getvalue()
        
        except ImportError:
            logger.error("openpyxl not installed, cannot export to Excel")
            return None
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return None
    
    def _export_multiple_to_csv(self, all_poll_data: List[Dict[str, Any]], options: ExportOptions) -> bytes:
        """Export multiple polls to CSV format."""
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        headers = ['Poll ID', 'Question', 'Option', 'Vote Count', 'Vote Type', 'Status']
        if options.include_metadata:
            headers.extend(['Team ID', 'Channel ID', 'Creator', 'Created At'])
        writer.writerow(headers)
        
        # Data rows for all polls
        for poll_data in all_poll_data:
            for option in poll_data['options']:
                row = [
                    poll_data['poll_id'],
                    poll_data['question'],
                    option['text'],
                    option['vote_count'],
                    poll_data['vote_type'],
                    poll_data['status']
                ]
                
                if options.include_metadata:
                    metadata = poll_data.get('metadata', {})
                    row.extend([
                        metadata.get('team_id', ''),
                        metadata.get('channel_id', ''),
                        metadata.get('creator_id', ''),
                        metadata.get('created_at', '')
                    ])
                
                writer.writerow(row)
        
        return output.getvalue().encode('utf-8')
    
    def _export_multiple_to_json(self, all_poll_data: List[Dict[str, Any]], options: ExportOptions) -> bytes:
        """Export multiple polls to JSON format."""
        export_data = {
            'exported_at': datetime.now().isoformat(),
            'export_options': {
                'include_voter_ids': options.include_voter_ids,
                'include_timestamps': options.include_timestamps,
                'include_analytics': options.include_analytics,
                'include_metadata': options.include_metadata,
                'anonymize_data': options.anonymize_data
            },
            'polls_count': len(all_poll_data),
            'polls_data': all_poll_data
        }
        
        return json.dumps(export_data, indent=2, ensure_ascii=False).encode('utf-8')
    
    def _export_multiple_to_excel(self, all_poll_data: List[Dict[str, Any]], options: ExportOptions) -> bytes:
        """Export multiple polls to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font
            
            workbook = openpyxl.Workbook()
            
            # Summary sheet with all polls
            summary_sheet = workbook.active
            summary_sheet.title = "All Polls Summary"
            
            # Headers
            summary_sheet['A1'] = "All Polls Summary"
            summary_sheet['A1'].font = Font(bold=True, size=14)
            
            # Column headers
            row = 3
            headers = ['Poll ID', 'Question', 'Option', 'Vote Count', 'Vote Type', 'Status']
            for col, header in enumerate(headers, 1):
                summary_sheet.cell(row=row, column=col, value=header).font = Font(bold=True)
            row += 1
            
            # Data for all polls
            for poll_data in all_poll_data:
                for option in poll_data['options']:
                    summary_sheet.cell(row=row, column=1, value=poll_data['poll_id'])
                    summary_sheet.cell(row=row, column=2, value=poll_data['question'])
                    summary_sheet.cell(row=row, column=3, value=option['text'])
                    summary_sheet.cell(row=row, column=4, value=option['vote_count'])
                    summary_sheet.cell(row=row, column=5, value=poll_data['vote_type'])
                    summary_sheet.cell(row=row, column=6, value=poll_data['status'])
                    row += 1
            
            # Create individual sheets for each poll
            for poll_data in all_poll_data:
                sheet_name = f"Poll_{poll_data['poll_id']}"
                poll_sheet = workbook.create_sheet(sheet_name)
                
                poll_sheet['A1'] = f"Poll {poll_data['poll_id']}: {poll_data['question']}"
                poll_sheet['A1'].font = Font(bold=True, size=12)
                
                # Options
                row = 3
                poll_sheet['A3'] = "Option"
                poll_sheet['B3'] = "Vote Count"
                poll_sheet['A3'].font = Font(bold=True)
                poll_sheet['B3'].font = Font(bold=True)
                row += 1
                
                for option in poll_data['options']:
                    poll_sheet.cell(row=row, column=1, value=option['text'])
                    poll_sheet.cell(row=row, column=2, value=option['vote_count'])
                    row += 1
            
            # Save to bytes
            output = BytesIO()
            workbook.save(output)
            return output.getvalue()
        
        except ImportError:
            logger.error("openpyxl not installed, cannot export to Excel")
            return None
        except Exception as e:
            logger.error(f"Error exporting multiple polls to Excel: {e}")
            return None

# Global exporter instance
poll_exporter = PollExporter()

# Utility functions
def export_poll_data(poll_id: int, format_type: str, include_voter_ids: bool = False, 
                    include_analytics: bool = True, anonymize: bool = True) -> Optional[bytes]:
    """Export a single poll."""
    options = ExportOptions(
        include_voter_ids=include_voter_ids,
        include_analytics=include_analytics,
        anonymize_data=anonymize
    )
    return poll_exporter.export_poll(poll_id, format_type, options)

def export_multiple_polls_data(poll_ids: List[int], format_type: str, include_analytics: bool = True, 
                              anonymize: bool = True) -> Optional[bytes]:
    """Export multiple polls."""
    options = ExportOptions(
        include_analytics=include_analytics,
        anonymize_data=anonymize
    )
    return poll_exporter.export_multiple_polls(poll_ids, format_type, options)

def get_supported_export_formats() -> List[str]:
    """Get list of supported export formats."""
    return poll_exporter.supported_formats