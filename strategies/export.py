"""
Export strategies implementing Strategy pattern.
Allows for extensible export formats following Open/Closed Principle.
"""

from abc import ABC, abstractmethod
from typing import Dict, Any, List, Optional
import json
import csv
import io
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExportStrategy(ABC):
    """Abstract base class for export strategies."""
    
    @abstractmethod
    def export(self, data: Dict[str, Any], options: Dict[str, Any] = None) -> bytes:
        """Export data to specific format."""
        pass
    
    @abstractmethod
    def get_format_name(self) -> str:
        """Get format name."""
        pass
    
    @abstractmethod
    def get_file_extension(self) -> str:
        """Get file extension for this format."""
        pass
    
    @abstractmethod
    def get_mime_type(self) -> str:
        """Get MIME type for this format."""
        pass


class CSVExportStrategy(ExportStrategy):
    """CSV export strategy."""
    
    def export(self, data: Dict[str, Any], options: Dict[str, Any] = None) -> bytes:
        """Export data to CSV format."""
        options = options or {}
        include_analytics = options.get('include_analytics', True)
        anonymize = options.get('anonymize', True)
        
        output = io.StringIO()
        
        # Determine if this is single poll or multiple polls
        if 'poll_data' in data:
            # Single poll export
            self._export_single_poll_csv(data, output, include_analytics, anonymize)
        elif 'polls' in data:
            # Multiple polls export
            self._export_multiple_polls_csv(data, output, include_analytics, anonymize)
        else:
            raise ValueError("Invalid data structure for CSV export")
        
        return output.getvalue().encode('utf-8')
    
    def _export_single_poll_csv(self, data: Dict[str, Any], output: io.StringIO, 
                               include_analytics: bool, anonymize: bool) -> None:
        """Export single poll to CSV."""
        poll_data = data['poll_data']
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['Poll Information'])
        writer.writerow(['Poll ID', poll_data.get('id', '')])
        writer.writerow(['Question', poll_data.get('question', '')])
        writer.writerow(['Vote Type', poll_data.get('vote_type', '')])
        writer.writerow(['Status', poll_data.get('status', '')])
        writer.writerow(['Created At', poll_data.get('created_at', '')])
        writer.writerow(['Creator', 'Anonymous' if anonymize else poll_data.get('creator_id', '')])
        writer.writerow([])
        
        # Options
        writer.writerow(['Options and Results'])
        writer.writerow(['Option', 'Votes', 'Percentage'])
        
        options = poll_data.get('options', [])
        total_votes = sum(opt.get('vote_count', 0) for opt in options)
        
        for option in options:
            votes = option.get('vote_count', 0)
            percentage = (votes / total_votes * 100) if total_votes > 0 else 0
            writer.writerow([
                option.get('text', ''),
                votes,
                f"{percentage:.1f}%"
            ])
        
        if include_analytics and 'analytics' in data:
            writer.writerow([])
            writer.writerow(['Analytics'])
            analytics = data['analytics']
            writer.writerow(['Total Votes', analytics.get('total_votes', 0)])
            writer.writerow(['Unique Voters', analytics.get('unique_voters', 0)])
            writer.writerow(['Participation Rate', f"{analytics.get('participation_rate', 0):.1f}%"])
    
    def _export_multiple_polls_csv(self, data: Dict[str, Any], output: io.StringIO,
                                  include_analytics: bool, anonymize: bool) -> None:
        """Export multiple polls to CSV."""
        writer = csv.writer(output)
        
        # Header row
        headers = ['Poll ID', 'Question', 'Vote Type', 'Status', 'Created At', 'Total Votes']
        if not anonymize:
            headers.append('Creator')
        if include_analytics:
            headers.extend(['Unique Voters', 'Participation Rate'])
        
        writer.writerow(headers)
        
        # Data rows
        for poll in data.get('polls', []):
            row = [
                poll.get('id', ''),
                poll.get('question', ''),
                poll.get('vote_type', ''),
                poll.get('status', ''),
                poll.get('created_at', ''),
                poll.get('total_votes', 0)
            ]
            
            if not anonymize:
                row.append(poll.get('creator_id', ''))
            
            if include_analytics:
                analytics = poll.get('analytics', {})
                row.extend([
                    analytics.get('unique_voters', 0),
                    f"{analytics.get('participation_rate', 0):.1f}%"
                ])
            
            writer.writerow(row)
    
    def get_format_name(self) -> str:
        return "CSV"
    
    def get_file_extension(self) -> str:
        return "csv"
    
    def get_mime_type(self) -> str:
        return "text/csv"


class JSONExportStrategy(ExportStrategy):
    """JSON export strategy."""
    
    def export(self, data: Dict[str, Any], options: Dict[str, Any] = None) -> bytes:
        """Export data to JSON format."""
        options = options or {}
        include_analytics = options.get('include_analytics', True)
        anonymize = options.get('anonymize', True)
        
        export_data = {
            'exported_at': datetime.now().isoformat(),
            'format': 'json',
            'options': {
                'include_analytics': include_analytics,
                'anonymize': anonymize
            }
        }
        
        # Process data based on structure
        if 'poll_data' in data:
            # Single poll export
            poll_data = data['poll_data'].copy()
            if anonymize:
                poll_data.pop('creator_id', None)
                poll_data.pop('user_votes', None)
            
            export_data['poll'] = poll_data
            
            if include_analytics and 'analytics' in data:
                export_data['analytics'] = data['analytics']
        
        elif 'polls' in data:
            # Multiple polls export
            polls = []
            for poll in data['polls']:
                poll_copy = poll.copy()
                if anonymize:
                    poll_copy.pop('creator_id', None)
                    poll_copy.pop('user_votes', None)
                polls.append(poll_copy)
            
            export_data['polls'] = polls
            export_data['total_polls'] = len(polls)
        
        return json.dumps(export_data, indent=2, default=str).encode('utf-8')
    
    def get_format_name(self) -> str:
        return "JSON"
    
    def get_file_extension(self) -> str:
        return "json"
    
    def get_mime_type(self) -> str:
        return "application/json"


class ExcelExportStrategy(ExportStrategy):
    """Excel export strategy."""
    
    def export(self, data: Dict[str, Any], options: Dict[str, Any] = None) -> bytes:
        """Export data to Excel format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")
        
        options = options or {}
        include_analytics = options.get('include_analytics', True)
        anonymize = options.get('anonymize', True)
        
        wb = Workbook()
        
        # Remove default worksheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        if 'poll_data' in data:
            # Single poll export
            self._export_single_poll_excel(data, wb, include_analytics, anonymize)
        elif 'polls' in data:
            # Multiple polls export
            self._export_multiple_polls_excel(data, wb, include_analytics, anonymize)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _export_single_poll_excel(self, data: Dict[str, Any], wb, include_analytics: bool, anonymize: bool) -> None:
        """Export single poll to Excel."""
        from openpyxl.styles import Font, PatternFill
        
        poll_data = data['poll_data']
        
        # Poll Info Sheet
        ws_info = wb.create_sheet("Poll Information")
        
        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Poll information
        ws_info['A1'] = "Property"
        ws_info['B1'] = "Value"
        ws_info['A1'].font = header_font
        ws_info['B1'].font = header_font
        ws_info['A1'].fill = header_fill
        ws_info['B1'].fill = header_fill
        
        info_rows = [
            ("Poll ID", poll_data.get('id', '')),
            ("Question", poll_data.get('question', '')),
            ("Vote Type", poll_data.get('vote_type', '')),
            ("Status", poll_data.get('status', '')),
            ("Created At", poll_data.get('created_at', '')),
            ("Creator", 'Anonymous' if anonymize else poll_data.get('creator_id', ''))
        ]
        
        for i, (prop, value) in enumerate(info_rows, 2):
            ws_info[f'A{i}'] = prop
            ws_info[f'B{i}'] = value
        
        # Options Sheet
        ws_options = wb.create_sheet("Options and Results")
        
        # Headers
        ws_options['A1'] = "Option"
        ws_options['B1'] = "Votes"
        ws_options['C1'] = "Percentage"
        
        for cell in ['A1', 'B1', 'C1']:
            ws_options[cell].font = header_font
            ws_options[cell].fill = header_fill
        
        # Options data
        options = poll_data.get('options', [])
        total_votes = sum(opt.get('vote_count', 0) for opt in options)
        
        for i, option in enumerate(options, 2):
            votes = option.get('vote_count', 0)
            percentage = (votes / total_votes * 100) if total_votes > 0 else 0
            
            ws_options[f'A{i}'] = option.get('text', '')
            ws_options[f'B{i}'] = votes
            ws_options[f'C{i}'] = f"{percentage:.1f}%"
        
        # Analytics sheet (if requested)
        if include_analytics and 'analytics' in data:
            ws_analytics = wb.create_sheet("Analytics")
            analytics = data['analytics']
            
            ws_analytics['A1'] = "Metric"
            ws_analytics['B1'] = "Value"
            ws_analytics['A1'].font = header_font
            ws_analytics['B1'].font = header_font
            ws_analytics['A1'].fill = header_fill
            ws_analytics['B1'].fill = header_fill
            
            analytics_rows = [
                ("Total Votes", analytics.get('total_votes', 0)),
                ("Unique Voters", analytics.get('unique_voters', 0)),
                ("Participation Rate", f"{analytics.get('participation_rate', 0):.1f}%"),
                ("Average Response Time", f"{analytics.get('avg_response_time', 0):.1f} minutes")
            ]
            
            for i, (metric, value) in enumerate(analytics_rows, 2):
                ws_analytics[f'A{i}'] = metric
                ws_analytics[f'B{i}'] = value
    
    def _export_multiple_polls_excel(self, data: Dict[str, Any], wb, include_analytics: bool, anonymize: bool) -> None:
        """Export multiple polls to Excel."""
        from openpyxl.styles import Font, PatternFill
        
        ws = wb.create_sheet("Polls Summary")
        
        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Headers
        headers = ['Poll ID', 'Question', 'Vote Type', 'Status', 'Created At', 'Total Votes']
        if not anonymize:
            headers.append('Creator')
        if include_analytics:
            headers.extend(['Unique Voters', 'Participation Rate'])
        
        for i, header in enumerate(headers, 1):
            cell = ws[f'{chr(64+i)}1']
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Data
        for row_idx, poll in enumerate(data.get('polls', []), 2):
            col_idx = 1
            
            # Basic poll data
            ws[f'{chr(64+col_idx)}{row_idx}'] = poll.get('id', '')
            col_idx += 1
            ws[f'{chr(64+col_idx)}{row_idx}'] = poll.get('question', '')
            col_idx += 1
            ws[f'{chr(64+col_idx)}{row_idx}'] = poll.get('vote_type', '')
            col_idx += 1
            ws[f'{chr(64+col_idx)}{row_idx}'] = poll.get('status', '')
            col_idx += 1
            ws[f'{chr(64+col_idx)}{row_idx}'] = poll.get('created_at', '')
            col_idx += 1
            ws[f'{chr(64+col_idx)}{row_idx}'] = poll.get('total_votes', 0)
            col_idx += 1
            
            if not anonymize:
                ws[f'{chr(64+col_idx)}{row_idx}'] = poll.get('creator_id', '')
                col_idx += 1
            
            if include_analytics:
                analytics = poll.get('analytics', {})
                ws[f'{chr(64+col_idx)}{row_idx}'] = analytics.get('unique_voters', 0)
                col_idx += 1
                ws[f'{chr(64+col_idx)}{row_idx}'] = f"{analytics.get('participation_rate', 0):.1f}%"
    
    def get_format_name(self) -> str:
        return "Excel"
    
    def get_file_extension(self) -> str:
        return "xlsx"
    
    def get_mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ExportContext:
    """Context class for managing export strategies."""
    
    def __init__(self):
        self.strategies: Dict[str, ExportStrategy] = {}
        
        # Register default strategies
        self.add_strategy(CSVExportStrategy())
        self.add_strategy(JSONExportStrategy())
        self.add_strategy(ExcelExportStrategy())
    
    def add_strategy(self, strategy: ExportStrategy) -> None:
        """Add an export strategy."""
        format_name = strategy.get_format_name().lower()
        self.strategies[format_name] = strategy
        logger.debug(f"Added export strategy: {format_name}")
    
    def remove_strategy(self, format_name: str) -> None:
        """Remove an export strategy."""
        format_name = format_name.lower()
        if format_name in self.strategies:
            del self.strategies[format_name]
            logger.debug(f"Removed export strategy: {format_name}")
    
    def export(self, data: Dict[str, Any], format_name: str, options: Dict[str, Any] = None) -> bytes:
        """Export data using specified format."""
        format_name = format_name.lower()
        
        if format_name not in self.strategies:
            raise ValueError(f"Unsupported export format: {format_name}")
        
        strategy = self.strategies[format_name]
        
        try:
            return strategy.export(data, options)
        except Exception as e:
            logger.error(f"Error in export strategy {format_name}: {e}")
            raise
    
    def get_supported_formats(self) -> List[Dict[str, str]]:
        """Get list of supported export formats."""
        return [
            {
                'name': strategy.get_format_name(),
                'extension': strategy.get_file_extension(),
                'mime_type': strategy.get_mime_type()
            }
            for strategy in self.strategies.values()
        ]
    
    def get_strategy(self, format_name: str) -> Optional[ExportStrategy]:
        """Get specific strategy by format name."""
        return self.strategies.get(format_name.lower())
    
    def has_format(self, format_name: str) -> bool:
        """Check if format is supported."""
        return format_name.lower() in self.strategies